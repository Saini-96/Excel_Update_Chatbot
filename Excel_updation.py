import openpyxl

'''
Only for understanding of the user:

For Rahul we have to pass this as an argument to the Excel_update function

update = {'Release': '12/01/2010',
         'User Story/Work Description': 'US00001:Analysis',
         'Appl': 'DEV',
         'Analysis': '10%',
         'Doc': '0%',
         'Coding': '50%',
         'Unit Testing': '0%',
         'Planned_Del': '12/04/2010',
         'Actual_Del': '12/06/2010',
         'Testing Support': 'No',
         'Planned_Imp': '12/15/2010',
         'Actual_Imp': '12/15/2010',
         'Remarks': 'Analysis in Progress'}


For Try we have to pass this as an argument to the Excel_update function
		 
update_2 = {'Release': '12/01/2010',
         'User Story/Work Description': 'US00002:Coding',
         'Appl': 'DEV',
         'Analysis': '100%',
         'Doc': '50%',
         'Coding': '50%',
         'Unit Testing': '30%',
         'Planned_Del': '12/04/2010',
         'Actual_Del': '12/06/2010',
         'Testing Support': 'Yes',
         'Planned_Imp': '12/15/2010',
         'Actual_Imp': '12/15/2010',
         'Remarks': 'Coding in Progress'}
'''


def Excel_update(user_name, Spreadsheet_name, update):
    
    wb = openpyxl.load_workbook(Spreadsheet_name)
    sheet = wb['Sheet1']
    row_start = 0

    for rowNum in range(1, sheet.max_row+1):
        User_Name = sheet.cell(row=rowNum, column=1).value
        if row_start == 0:
            row_start = rowNum
        if User_Name == user_name:
            for colNum in range(2, sheet.max_column+1):
                Column_Name = sheet.cell(row=1, column=colNum).value
                if Column_Name in ('Delivery to QA' , 'Implementation Date'):
                    row_update = row_start + 1
                    Column_Name = sheet.cell(row=row_update, column=colNum).value
                    if Column_Name in update:
                        sheet.cell(row=rowNum, column=colNum).value = update[Column_Name]     
                        colNum = colNum + 1
                        Column_Name = sheet.cell(row=row_update, column=colNum).value
                        if Column_Name in update:
                            sheet.cell(row=rowNum, column=colNum).value = update[Column_Name]
                if Column_Name in update:
                    sheet.cell(row=rowNum, column=colNum).value = update[Column_Name]
    print("\nExcel Updatation Done Successfully....\n")
             
    wb.save(Spreadsheet_name)


'''
This function will be called for updating status in Excel sheet through Chatbot interface

Excel_update("Try", "Sample.xlsx", update = {'Release': '12/01/2010',
         'User Story/Work Description': 'US00002:Coding',
         'Appl': 'DEV',
         'Analysis': '100%',
         'Doc': '50%',
         'Coding': '50%',
         'Unit Testing': '30%',
         'Planned_Del': '12/04/2010',
         'Actual_Del': '12/06/2010',
         'Testing Support': 'Yes',
         'Planned_Imp': '12/15/2010',
         'Actual_Imp': '12/15/2010',
         'Remarks': 'Coding in Progress'})
''' 